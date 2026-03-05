#!/usr/bin/env python3
"""
Sync between Excel and the web app's JSON format.

Usage:
    # Export Excel → JSON (to import into web app)
    python sync_excel_json.py export --excel contractors.xlsx --json contractors.json

    # Import JSON → Excel (from web app export)
    python sync_excel_json.py import --json contractors.json --excel contractors.xlsx
"""

import json
import sys
import argparse
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    sys.exit("Missing dependency: pip install openpyxl")


COLUMNS = [
    "Name", "Contact", "Phone", "Email", "Website",
    "Address", "Services", "Industry", "Rating", "Notes",
    "Tags", "Photo File", "Date Added"
]

FIELD_MAP = {
    "Name": "name", "Contact": "contact", "Phone": "phone",
    "Email": "email", "Website": "website", "Address": "address",
    "Services": "services", "Industry": "industry", "Rating": "rating",
    "Notes": "notes", "Tags": "tags"
}


def excel_to_json(excel_path, json_path):
    """Export Excel rows to JSON format compatible with the web app."""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    contractors = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        entry = {}
        for i, val in enumerate(row):
            if i < len(headers) and headers[i] in FIELD_MAP:
                field = FIELD_MAP[headers[i]]
                if field == "tags" and isinstance(val, str):
                    entry[field] = [t.strip() for t in val.split(",") if t.strip()]
                elif field == "rating":
                    entry[field] = int(val) if val else 0
                else:
                    entry[field] = str(val).strip() if val else ""
        if entry.get("name"):
            contractors.append(entry)

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(contractors, f, indent=2, ensure_ascii=False)

    print(f"Exported {len(contractors)} contractors to {json_path}")
    print("You can import this JSON into the web app using the menu → Import JSON")


def json_to_excel(json_path, excel_path):
    """Import JSON from the web app into Excel."""
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contractors"

    # Headers
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill

    widths = [30, 20, 20, 25, 25, 35, 40, 20, 8, 35, 30, 25, 15]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Rows
    for entry in data:
        tags = entry.get("tags", [])
        if isinstance(tags, list):
            tags = ", ".join(tags)
        row_data = [
            entry.get("name", ""),
            entry.get("contact", ""),
            entry.get("phone", ""),
            entry.get("email", ""),
            entry.get("website", ""),
            entry.get("address", ""),
            entry.get("services", ""),
            entry.get("industry", ""),
            entry.get("rating", 0),
            entry.get("notes", ""),
            tags,
            "",  # Photo file
            entry.get("createdAt", datetime.now().strftime("%Y-%m-%d %H:%M"))
        ]
        ws.append(row_data)

    ws.auto_filter.ref = ws.dimensions
    wb.save(excel_path)
    print(f"Imported {len(data)} contractors to {excel_path}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Sync between Excel and web app JSON")
    sub = parser.add_subparsers(dest="command")

    exp = sub.add_parser("export", help="Excel → JSON")
    exp.add_argument("--excel", "-e", required=True)
    exp.add_argument("--json", "-j", required=True)

    imp = sub.add_parser("import", help="JSON → Excel")
    imp.add_argument("--json", "-j", required=True)
    imp.add_argument("--excel", "-e", required=True)

    args = parser.parse_args()
    if args.command == "export":
        excel_to_json(args.excel, args.json)
    elif args.command == "import":
        json_to_excel(args.json, args.excel)
    else:
        parser.print_help()
