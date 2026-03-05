#!/usr/bin/env python3
"""
Auto-Scan Contractor Photos → Excel
Watches a folder for new business card / truck signage photos,
runs OCR, parses contractor info, and updates an Excel spreadsheet.

Schedule this with Windows Task Scheduler or cron to run daily.

Usage:
    python auto_scan.py
    python auto_scan.py --folder "C:\path\to\photos" --excel "C:\path\to\contractors.xlsx"
"""

import os
import re
import sys
import shutil
import argparse
import logging
from pathlib import Path
from datetime import datetime

# ── Dependencies ──
try:
    from PIL import Image
except ImportError:
    sys.exit("Missing dependency: pip install Pillow")
try:
    import pytesseract
except ImportError:
    sys.exit("Missing dependency: pip install pytesseract\nAlso install Tesseract OCR: https://github.com/tesseract-ocr/tesseract")
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    sys.exit("Missing dependency: pip install openpyxl")


# ═══════════════════════════════════════════════════════════════════
# Configuration
# ═══════════════════════════════════════════════════════════════════

DEFAULT_FOLDER = r"C:\Users\shaya\OneDrive - Oak builders llc\Oak\Bussiness Cards"
DEFAULT_EXCEL = os.path.join(DEFAULT_FOLDER, "contractors.xlsx")
PROCESSED_SUBFOLDER = "processed"
LOG_FILE = "auto_scan.log"
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".bmp", ".tiff", ".tif", ".webp", ".heic"}

# Excel column layout matching the web app's data structure
COLUMNS = [
    "Name", "Contact", "Phone", "Email", "Website",
    "Address", "Services", "Industry", "Rating", "Notes",
    "Tags", "Photo File", "Date Added"
]

# Industry keyword mapping (mirrors the web app's INDUSTRY_KEYWORDS)
INDUSTRY_KEYWORDS = {
    "Plumbing": ["plumbing", "plumber", "drain", "sewer", "water heater", "pipe fitting", "faucet"],
    "HVAC": ["hvac", "heating", "cooling", "air conditioning", "furnace", "ductwork", "heat pump"],
    "Electrical": ["electrical", "electrician", "wiring", "panel", "circuit", "lighting"],
    "General Contractor": ["general contractor", "remodeling", "renovation", "home improvement", "additions", "construction"],
    "Roofing": ["roofing", "roof ", "shingle", "gutter", "flashing"],
    "Painting": ["painting", "painter", "interior painting", "exterior painting", "staining"],
    "Masonry": ["masonry", "brick", "concrete", "stone", "stucco", "block"],
    "Flooring": ["flooring", "hardwood", "tile", "carpet", "lvp", "vinyl plank", "laminate"],
    "Waterproofing": ["waterproofing", "basement waterproof", "foundation repair", "french drain"],
    "Tree Service": ["tree service", "tree removal", "stump", "arborist", "tree trimming"],
    "Hauling": ["hauling", "junk removal", "debris", "dumpster", "cleanout"],
    "Landscaping": ["landscaping", "lawn", "garden", "hardscape", "patio", "retaining wall", "irrigation"],
    "Drywall": ["drywall", "sheetrock", "plastering", "texture"],
    "Siding": ["siding", "vinyl siding", "hardie", "fiber cement"],
    "Windows & Doors": ["window", "door", "glass", "replacement window"],
    "Demolition": ["demolition", "demo", "teardown", "abatement"],
    "Framing": ["framing", "rough carpentry", "structural"],
    "Finish Carpentry": ["trim", "molding", "cabinet", "finish carpentry", "millwork"],
    "Insulation": ["insulation", "spray foam", "blown-in", "batt"],
    "Excavation": ["excavation", "grading", "earthwork", "trenching", "site work"],
    "Pest Control": ["pest control", "exterminator", "termite", "rodent"],
    "Cleaning": ["cleaning", "janitorial", "pressure wash", "power wash"],
}


# ═══════════════════════════════════════════════════════════════════
# Logging
# ═══════════════════════════════════════════════════════════════════

def setup_logging(folder):
    log_path = os.path.join(folder, LOG_FILE)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[
            logging.FileHandler(log_path, encoding="utf-8"),
            logging.StreamHandler()
        ]
    )


# ═══════════════════════════════════════════════════════════════════
# OCR + Parsing (mirrors the web app's parseBusinessCard logic)
# ═══════════════════════════════════════════════════════════════════

def extract_text(image_path):
    """Run Tesseract OCR on an image file."""
    img = Image.open(image_path)
    # Resize large images for faster OCR
    max_dim = 2000
    if max(img.size) > max_dim:
        ratio = max_dim / max(img.size)
        img = img.resize((int(img.width * ratio), int(img.height * ratio)), Image.LANCZOS)
    text = pytesseract.image_to_string(img)
    return text


def parse_business_card(text):
    """Parse OCR text into structured contractor fields."""
    lines = [l.strip() for l in text.split("\n") if l.strip()]
    fields = {
        "name": "", "contact": "", "phone": "", "email": "",
        "website": "", "address": "", "services": "", "industry": "",
        "rating": 0, "notes": "", "tags": ""
    }
    used_lines = set()

    # ── Phone numbers (capture ALL) ──
    phone_re = re.compile(
        r'(?:\+?1[-.\s]?)?\(?\d{3}\)?[-.\s]?\d{3}[-.\s]?\d{4}'
        r'|1[-.\s]?8[0-9]{2}[-.\s]?\d{3}[-.\s]?\d{4}'
    )
    all_phones = []
    for i, line in enumerate(lines):
        for m in phone_re.finditer(line):
            all_phones.append(m.group().strip())
            used_lines.add(i)
    fields["phone"] = ", ".join(all_phones)

    # ── Email ──
    email_re = re.compile(r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}')
    for i, line in enumerate(lines):
        m = email_re.search(line)
        if m:
            fields["email"] = m.group().lower()
            used_lines.add(i)
            break

    # ── Website ──
    web_patterns = [
        re.compile(r'(?:www\.)[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}(?:/\S*)?', re.I),
        re.compile(r'(?:https?://)[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}(?:/\S*)?', re.I),
        re.compile(r'[a-zA-Z0-9][-a-zA-Z0-9]*\.(?:com|net|org|io|co|biz|info|us)(?:/\S*)?', re.I),
    ]
    for i, line in enumerate(lines):
        if "@" in line:
            continue
        for pat in web_patterns:
            m = pat.search(line)
            if m:
                fields["website"] = m.group()
                used_lines.add(i)
                break
        if fields["website"]:
            break

    # ── Address ──
    addr_re = re.compile(
        r'\d+\s+[\w\s]+(?:St|Street|Ave|Avenue|Blvd|Boulevard|Dr|Drive|Rd|Road|Ln|Lane|Way|Ct|Court|Pl|Place)'
        r'[^,]*,?\s*[A-Z]{2}\s*\d{5}',
        re.I
    )
    for i, line in enumerate(lines):
        m = addr_re.search(line)
        if m:
            fields["address"] = m.group().strip()
            used_lines.add(i)
            break
    # Fallback: state + zip
    if not fields["address"]:
        state_zip = re.compile(r'[A-Z]{2}\s+\d{5}(?:-\d{4})?')
        for i, line in enumerate(lines):
            if state_zip.search(line):
                fields["address"] = line.strip()
                used_lines.add(i)
                break

    # ── Credential lines → notes ──
    credential_re = re.compile(
        r'licens|bond|insur|certified|family.owned|established|since \d{4}|veteran|warranty|guaranteed',
        re.I
    )
    note_parts = []
    for i, line in enumerate(lines):
        if i not in used_lines and credential_re.search(line):
            note_parts.append(line)
            used_lines.add(i)
    fields["notes"] = "; ".join(note_parts)

    # ── Services (bullet/pipe/comma-separated lists) ──
    service_separators = re.compile(r'[•|/\\]')
    service_parts = []
    for i, line in enumerate(lines):
        if i in used_lines:
            continue
        if service_separators.search(line) or (line.count(",") >= 2):
            normalized = service_separators.sub(",", line)
            items = [s.strip() for s in normalized.split(",") if len(s.strip()) > 1]
            service_parts.extend(items)
            used_lines.add(i)
        elif line.isupper() and len(line) > 3 and not phone_re.search(line):
            # ALL-CAPS lines are often service descriptions on trucks
            service_parts.append(line.title())
            used_lines.add(i)
    fields["services"] = ", ".join(service_parts)

    # ── Company name (score remaining lines) ──
    candidates = [(i, line) for i, line in enumerate(lines) if i not in used_lines]
    if candidates:
        def name_score(idx, line):
            score = 0
            if idx < 3:
                score += (3 - idx) * 10
            caps = sum(1 for w in line.split() if w and w[0].isupper())
            score += caps * 2
            if len(line) > 50:
                score -= 10
            if line.count(",") > 3:
                score -= 15
            if credential_re.search(line):
                score -= 20
            # Boost lines with construction-related words
            construction_words = ["construction", "builders", "building", "contractor",
                                  "services", "company", "solutions", "enterprises", "group", "llc", "inc"]
            lower = line.lower()
            for w in construction_words:
                if w in lower:
                    score += 5
            return score

        best = max(candidates, key=lambda c: name_score(c[0], c[1]))
        fields["name"] = best[1].strip()[:200]
        used_lines.add(best[0])

    # ── Industry auto-detection ──
    combined = (fields["services"] + " " + fields["name"] + " " + fields["notes"]).lower()
    best_industry = ""
    best_score = 0
    for industry, keywords in INDUSTRY_KEYWORDS.items():
        score = sum(1 for kw in keywords if kw in combined)
        if score > best_score:
            best_score = score
            best_industry = industry
    if best_score > 0:
        fields["industry"] = best_industry

    # ── Tags from services ──
    if fields["services"]:
        tag_items = [s.strip() for s in re.split(r'[,;]', fields["services"]) if s.strip()]
        fields["tags"] = ", ".join(tag_items[:8])

    return fields


# ═══════════════════════════════════════════════════════════════════
# Excel Operations
# ═══════════════════════════════════════════════════════════════════

def create_excel(path):
    """Create a new Excel file with headers and formatting."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contractors"

    # Header styling
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        bottom=Side(style="thin", color="E2E5EA")
    )

    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Column widths
    widths = [30, 20, 20, 25, 25, 35, 40, 20, 8, 35, 30, 25, 15]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws.auto_filter.ref = ws.dimensions
    wb.save(path)
    return wb


def load_or_create_excel(path):
    """Load existing Excel or create a new one."""
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        # Verify headers match (check first column)
        if ws.cell(row=1, column=1).value != COLUMNS[0]:
            logging.warning("Excel headers don't match expected format. Adding a new sheet.")
            ws = wb.create_sheet("Contractors (Auto)")
            for col_idx, col_name in enumerate(COLUMNS, 1):
                ws.cell(row=1, column=col_idx, value=col_name)
        return wb, ws
    else:
        wb = create_excel(path)
        return wb, wb.active


def get_existing_names(ws):
    """Get set of existing company names (lowercase) to detect duplicates."""
    names = set()
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0]:
            names.add(row[0].strip().lower())
    return names


def normalize_name(name):
    """Normalize company name for duplicate comparison."""
    name = name.lower()
    # Strip common suffixes
    name = re.sub(r'\b(llc|inc|corp|co|ltd|company|services?|solutions?|group)\b', '', name)
    name = re.sub(r'[^a-z0-9]', '', name)
    return name


def is_duplicate(new_name, existing_names_normalized):
    """Check if a name is a duplicate using fuzzy matching."""
    new_norm = normalize_name(new_name)
    if len(new_norm) < 2:
        return False
    for existing in existing_names_normalized:
        # Substring match
        if new_norm in existing or existing in new_norm:
            return True
        # Simple similarity: if 80%+ characters overlap
        if len(new_norm) > 3 and len(existing) > 3:
            max_len = max(len(new_norm), len(existing))
            common = sum(1 for a, b in zip(new_norm, existing) if a == b)
            if common / max_len > 0.75:
                return True
    return False


def add_row(ws, fields, photo_filename):
    """Append a new contractor row to the worksheet."""
    next_row = ws.max_row + 1
    row_data = [
        fields.get("name", ""),
        fields.get("contact", ""),
        fields.get("phone", ""),
        fields.get("email", ""),
        fields.get("website", ""),
        fields.get("address", ""),
        fields.get("services", ""),
        fields.get("industry", ""),
        fields.get("rating", 0),
        fields.get("notes", ""),
        fields.get("tags", ""),
        photo_filename,
        datetime.now().strftime("%Y-%m-%d %H:%M")
    ]
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=next_row, column=col_idx, value=value)
        cell.alignment = Alignment(wrap_text=True, vertical="top")


# ═══════════════════════════════════════════════════════════════════
# Main Processing
# ═══════════════════════════════════════════════════════════════════

def find_new_images(folder, processed_folder):
    """Find image files that haven't been processed yet."""
    processed_names = set()
    if os.path.exists(processed_folder):
        processed_names = {f.lower() for f in os.listdir(processed_folder)}

    images = []
    for f in os.listdir(folder):
        if f.lower().startswith("."):
            continue
        ext = Path(f).suffix.lower()
        if ext in IMAGE_EXTENSIONS and f.lower() not in processed_names:
            full_path = os.path.join(folder, f)
            if os.path.isfile(full_path):
                images.append(full_path)
    return sorted(images)


def process_folder(folder, excel_path):
    """Main entry point: scan folder for new photos, OCR them, update Excel."""
    setup_logging(folder)
    logging.info("=" * 60)
    logging.info("Starting auto-scan run")
    logging.info(f"Folder: {folder}")
    logging.info(f"Excel:  {excel_path}")

    processed_folder = os.path.join(folder, PROCESSED_SUBFOLDER)
    os.makedirs(processed_folder, exist_ok=True)

    # Find new images
    new_images = find_new_images(folder, processed_folder)
    if not new_images:
        logging.info("No new images found. Exiting.")
        return 0

    logging.info(f"Found {len(new_images)} new image(s) to process")

    # Load Excel
    wb, ws = load_or_create_excel(excel_path)

    # Get existing names for duplicate detection
    existing_names = get_existing_names(ws)
    existing_normalized = {normalize_name(n) for n in existing_names}

    added = 0
    skipped_dupes = 0
    errors = 0

    for img_path in new_images:
        filename = os.path.basename(img_path)
        logging.info(f"Processing: {filename}")

        try:
            # OCR
            text = extract_text(img_path)
            if not text.strip():
                logging.warning(f"  No text detected in {filename}")
                errors += 1
                # Still move to processed to avoid re-scanning
                shutil.move(img_path, os.path.join(processed_folder, filename))
                continue

            logging.info(f"  OCR text: {text[:100]}...")

            # Parse
            fields = parse_business_card(text)

            if not fields["name"]:
                logging.warning(f"  Could not detect company name in {filename}")
                fields["name"] = f"Unknown ({filename})"

            # Duplicate check
            if is_duplicate(fields["name"], existing_normalized):
                logging.info(f"  DUPLICATE detected: '{fields['name']}' - skipping")
                skipped_dupes += 1
                # Move to processed but note it was a dupe
                fields["notes"] = (fields.get("notes", "") + " [DUPLICATE - review needed]").strip()
                # Still add it but flag it so user can review
                add_row(ws, fields, filename)
                added += 1
            else:
                # Add new entry
                add_row(ws, fields, filename)
                added += 1
                existing_normalized.add(normalize_name(fields["name"]))

            logging.info(f"  Added: {fields['name']} | {fields['phone']} | {fields['industry']}")

            # Move photo to processed
            shutil.move(img_path, os.path.join(processed_folder, filename))

        except Exception as e:
            logging.error(f"  Error processing {filename}: {e}")
            errors += 1

    # Save Excel
    try:
        wb.save(excel_path)
        logging.info(f"Excel saved: {excel_path}")
    except PermissionError:
        # Excel file might be open - save to alternate name
        alt_path = excel_path.replace(".xlsx", f"_update_{datetime.now():%Y%m%d_%H%M}.xlsx")
        wb.save(alt_path)
        logging.warning(f"Excel was open. Saved to: {alt_path}")

    logging.info(f"Done! Added: {added}, Duplicates flagged: {skipped_dupes}, Errors: {errors}")
    return added


# ═══════════════════════════════════════════════════════════════════
# CLI Entry Point
# ═══════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Auto-scan contractor photos and update Excel database"
    )
    parser.add_argument(
        "--folder", "-f",
        default=DEFAULT_FOLDER,
        help=f"Folder to scan for photos (default: {DEFAULT_FOLDER})"
    )
    parser.add_argument(
        "--excel", "-e",
        default=DEFAULT_EXCEL,
        help=f"Excel file path (default: {DEFAULT_EXCEL})"
    )
    args = parser.parse_args()

    if not os.path.isdir(args.folder):
        print(f"Error: Folder not found: {args.folder}")
        print("Create the folder or specify a different path with --folder")
        sys.exit(1)

    count = process_folder(args.folder, args.excel)
    print(f"\nProcessed {count} new contractor(s).")
